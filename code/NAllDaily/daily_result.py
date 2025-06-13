from lark_oapi.client import config
import win32com.client as win32
import time
from PIL import ImageGrab
import send_as_bot
import openpyxl
from openpyxl.utils import get_column_letter
import tool
import handle_sql
from datetime import datetime

def export_daily_result():
    today = datetime.today().strftime("%Y%m%d")
    today = '20250612'
    config_data = tool.read_ftp_config()
    product_data = []
    account_data = []

    if config_data:
        xlsx_config = config_data['xlsx_config']
        print(xlsx_config)

        # for product in config_data['products']:
        #     product_name = product['name']
        #     product_sql = handle_sql.select_product(date, product_name)
        #     print('产品', product_sql)
        #     product_data.extend(product_sql)

        #     fund_accounts = []
        #     for code in product['codes']:
        #         fund_accounts += code['fund_account']

        #     accounts = []
        #     for fund_account in fund_accounts:
        #         accounts += handle_sql.select_account(date, fund_account)

        #     print('账号', accounts)
        #     print('\n')
        #     account_data.extend(accounts)

        data_fields = ['netvalueest', 'totalassets', 'marketvalue', 'cash', 'position', 'dailyper']

        # 处理山西证券数据
        shanxi_product_sql = handle_sql.select_product(today, '山西证券')
        for field in data_fields:
            config_key = f'ShanXi_{field}'
            value = shanxi_product_sql[0][field]
            if field == 'dailyper':
                color = 'FF0000' if float(value.strip('%')) > 0 else '00FF00'
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value,
                    color
                )
            else:
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value
                )

        # 处理尊享2号数据
        zunxiang_product_sql = handle_sql.select_product(today, '尊享2号')
        for field in data_fields:
            config_key = f'ZunXiang_{field}'
            value = zunxiang_product_sql[0][field]
            if field == 'dailyper':
                color = 'FF0000' if float(value.strip('%')) > 0 else '00FF00'
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value,
                    color
                )
            else:
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value
                )

        # 处理九章量化数据
        jiuzhang_product_sql = handle_sql.select_product(today, '九章量化')
        for field in data_fields:
            config_key = f'JiuZhang_{field}'
            value = jiuzhang_product_sql[0][field]
            if field == 'dailyper':
                color = 'FF0000' if float(value.strip('%')) > 0 else '00FF00'
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value,
                    color
                )
            else:
                insert_data_to_excel(
                    xlsx_config[config_key]['row'],
                    xlsx_config[config_key]['col'],
                    value
                )

        # 处理账号数据
        data_fields = ['totalassets', 'marketvalue', 'cash', 'position', 'dailyper']
        jiuzhang_product_sql = handle_sql.select_product(today, '九章量化')
        accounts = [
            (190900011119, 'JiuZhang_GX'),
            (9220717, 'JiuZhang_GT'),
            (109157018941, 'JiuZhang_ZT')
        ]

        for account, prefix in accounts:
            account_sql = handle_sql.select_account(today, account)
            for field in data_fields:
                config_key = f'{prefix}_{field}'
                value = account_sql[0][field]
                if field == 'dailyper':
                    color = 'FF0000' if float(value.strip('%')) > 0 else '00FF00'
                    insert_data_to_excel(
                        xlsx_config[config_key]['row'],
                        xlsx_config[config_key]['col'],
                        value,
                        color
                    )
                else:
                    insert_data_to_excel(
                        xlsx_config[config_key]['row'],
                        xlsx_config[config_key]['col'],
                        value
                    )

            position_t = account_sql[0]['totalassets'] / jiuzhang_product_sql[0]['totalassets']
            position = account_sql[0]['position']
            insert_data_to_excel(
                xlsx_config[f'{prefix}_position']['row'],
                xlsx_config[f'{prefix}_position']['col'],
                position + f'({position_t * 100:.2f}%)'
            )

        insert_data_to_excel(xlsx_config['date']['row'], xlsx_config['date']['col'], today)
        time.sleep(2)
        capture_excel_screenshot()

def capture_excel_screenshot():
    """
    根据绝对路径打开 Excel 表格，全屏显示，然后根据坐标点截取图片到指定路径。

    :param file_path: Excel 文件的绝对路径
    :param coordinates: 截图的坐标点，格式为 (left, top, right, bottom)
    :param output_path: 截图保存的路径
    :return: 操作成功返回 True，失败返回 False
    """
    # 启动 Excel 应用程序
    excel = win32.gencache.EnsureDispatch('Excel.Application')
    # 使 Excel 可见
    excel.Visible = True
    try:
        config_data = tool.read_ftp_config()
        xlsx_path = config_data['xlsx_path']
        output_path = config_data['output_path']
        coordinate = config_data['coordinate']
        coordinates = (coordinate['left'], coordinate['top'], coordinate['right'], coordinate['bottom'])

        # 打开工作簿
        workbook = excel.Workbooks.Open(xlsx_path)
        # 全屏显示 Excel 窗口
        excel.Application.WindowState = -4137  # xlMaximized

        # 等待 Excel 窗口完全加载
        time.sleep(3)

        # 进行截图
        screenshot = ImageGrab.grab(bbox=coordinates)
        # 保存截图
        screenshot.save(output_path)

        time.sleep(3)
        send_as_bot.send_group_message(config_data['feishu_group_name'], True, output_path)
    except Exception as e:
        print(f"操作过程中出现错误: {e}")
    finally:
        # 关闭工作簿
        if 'workbook' in locals():
            workbook.Close(SaveChanges=False)
        # 退出 Excel 应用程序
        excel.Quit()

def insert_data_to_excel(row, col, data, color=None):
    """
    向指定 Excel 文件的指定行列插入数据，支持合并单元格的情况和设置字体颜色
    
    :param row: 要插入数据的行号（从 1 开始）
    :param col: 要插入数据的列号（从 1 开始）
    :param data: 要插入的数据
    :param color: 可选，字体颜色(十六进制格式，如'FF0000'表示红色)
    :return: 操作成功返回 True，失败返回 False
    """
    try:
        config_data = tool.read_ftp_config()
        xlsx_path = config_data['xlsx_path']
        workbook = openpyxl.load_workbook(xlsx_path)
        sheet = workbook.active

        # 检查指定位置是否在合并单元格内
        for merged_range in sheet.merged_cells.ranges:
            min_col, min_row, max_col, max_row = merged_range.bounds
            if min_row <= row <= max_row and min_col <= col <= max_col:
                cell = sheet.cell(row=min_row, column=min_col)
                cell.value = data
                if color:
                    # 保留原有字体样式，只修改颜色
                    original_font = cell.font
                    cell.font = openpyxl.styles.Font(
                        name=original_font.name,
                        size=original_font.size,
                        bold=original_font.bold,
                        italic=original_font.italic,
                        underline=original_font.underline,
                        strike=original_font.strike,
                        color=color
                    )
                break
        else:
            cell = sheet.cell(row=row, column=col)
            cell.value = data
            if color:
                original_font = cell.font
                cell.font = openpyxl.styles.Font(
                    name=original_font.name,
                    size=original_font.size,
                    bold=original_font.bold,
                    italic=original_font.italic,
                    underline=original_font.underline,
                    strike=original_font.strike,
                    color=color
                )

        workbook.save(xlsx_path)
        return True
    except Exception as e:
        print(f"插入数据时出现错误: {e}")
        return False

# 使用示例
if __name__ == "__main__":
    export_daily_result()