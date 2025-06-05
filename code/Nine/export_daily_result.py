import csv
import chardet
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.styles import Alignment


def detect_encoding(file_path):
    """检测文件编码"""
    with open(file_path, 'rb') as f:
        raw_data = f.read()
        result = chardet.detect(raw_data)
        return result['encoding']


def get_data_by_date(file_path, target_date):
    """根据指定日期从文件中获取数据"""
    encoding = detect_encoding(file_path)
    data = []
    with open(file_path, 'r', encoding=encoding) as csvfile:
        reader = csv.DictReader(csvfile)
        for row in reader:
            if row.get('时间') == target_date:
                data.append(row)
    return data


def get_net_value_from_excel(file_path, target_date):
    """从 Excel 文件中根据指定日期获取资产份额净值"""
    df = pd.read_excel(file_path)
    # 假设日期列名为 '日期'，将日期列转换为字符串格式，格式为 'YYYYMMDD'
    df['日期'] = pd.to_datetime(df['日期']).dt.strftime('%Y%m%d')
    filtered_df = df[df['日期'] == target_date]
    if not filtered_df.empty:
        return filtered_df['资产份额净值(元)'].values[0]
    return None


def get_daily_return(file_path, target_date):
    """
    从指定路径的 CSV 文件中根据指定日期获取当日日涨跌幅。
    :param file_path: 存储日涨跌幅数据的 CSV 文件路径
    :param target_date: 目标日期，格式为 'YYYYMMDD'
    :return: 若找到匹配日期，返回对应的日涨跌幅；未找到则返回 0
    """
    try:
        encoding = detect_encoding(file_path)
        with open(file_path, 'r', encoding=encoding) as csvfile:
            reader = csv.DictReader(csvfile)
            for row in reader:
                # 假设日期列名为 '日期'，日涨跌幅列名为 '日涨跌幅'
                if row.get('时间') == target_date:
                    try:
                        # 获取日涨跌幅数据
                        daily_return_str = row.get('日涨跌幅', '0')
                        # 去掉百分号并转换为浮点数，再除以 100
                        if daily_return_str.endswith('%'):
                            daily_return = float(daily_return_str.rstrip('%')) / 100
                        else:
                            daily_return = float(daily_return_str)
                        return daily_return
                    except ValueError:
                        print(f"无法将日涨跌幅 {row.get('日涨跌幅')} 转换为浮点数，使用默认值 0。")
                        return 0
    except FileNotFoundError:
        print(f"未找到日涨跌幅文件: {file_path}")
    return 0

def export_data(target_date, output_file):
    """导出指定日期的数据到新表格"""
    input_files = [
        r'c:\Users\Top\Desktop\九章量化\part-result.csv',
        r'c:\Users\Top\Desktop\九章量化\all-result.csv'
    ]
    all_data = []
    fieldnames = []
    seen_fields = set()

    # 按顺序收集所有列名
    for file in input_files:
        encoding = detect_encoding(file)
        with open(file, 'r', encoding=encoding) as csvfile:
            reader = csv.DictReader(csvfile)
            if reader.fieldnames:
                for field in reader.fieldnames:
                    if field not in seen_fields:
                        fieldnames.append(field)
                        seen_fields.add(field)

    # 从每个输入文件中获取指定日期的数据
    for file in input_files:
        all_data.extend(get_data_by_date(file, target_date))

    if not all_data:
        print(f"未找到 {target_date} 的数据。")
        return

    # 从 Excel 文件中获取净值
    excel_file_path = r'c:\Users\Top\Desktop\FuYing\code\Nine\ftp\富赢九章量化1号私募证券投资基金_SXA689_基金每日净值表.xls'
    if not os.path.exists(excel_file_path):
        print(f"未找到 Excel 文件: {excel_file_path}，请检查文件路径和文件是否存在。")
        return

    # 假设日涨跌幅文件路径
    daily_return_file = input_files[1]
    daily_return = get_daily_return(daily_return_file, target_date)
    print(f"日涨跌幅: {daily_return}")

    net_value = get_net_value_from_excel(excel_file_path, target_date)
    net_value = 1.9362
    if net_value is not None:
        net_value = net_value * (1 + daily_return)

    if net_value is not None:
        new_row = {}
        if '产品' in fieldnames:
            new_row['产品'] = '九章量化' + target_date
        if '总资产' in fieldnames:
            new_row['总资产'] = '净值'
        if '市值' in fieldnames:
            new_row['市值'] = round(net_value, 4)
        if new_row:
            all_data.append(new_row)

    # 过滤掉时间列
    new_fieldnames = [field for field in fieldnames if field != '时间']
    filtered_data = []
    for row in all_data:
        new_row = {key: value for key, value in row.items() if key != '时间'}
        filtered_data.append(new_row)

    # 创建一个新的 Excel 工作簿
    wb = Workbook()
    ws = wb.active

    # 写入表头
    ws.append(new_fieldnames)

    # 写入数据
    for row in filtered_data:
        ws.append([row.get(field, '') for field in new_fieldnames])

    # 查找 '产品' 列的索引
    if '产品' in new_fieldnames:
        product_col_idx = new_fieldnames.index('产品') + 1
        start_row = 2
        current_value = ws.cell(row=start_row, column=product_col_idx).value
        for row_idx in range(3, ws.max_row + 1):
            cell_value = ws.cell(row=row_idx, column=product_col_idx).value
            if cell_value != current_value:
                if start_row < row_idx - 1:
                    ws.merge_cells(start_row=start_row, start_column=product_col_idx, end_row=row_idx - 1, end_column=product_col_idx)
                    merged_cell = ws.cell(row=start_row, column=product_col_idx)
                    merged_cell.alignment = Alignment(vertical='center', horizontal='center')
                start_row = row_idx
                current_value = cell_value

        # 处理最后一组相同内容的单元格
        if start_row < ws.max_row:
            ws.merge_cells(start_row=start_row, start_column=product_col_idx, end_row=ws.max_row, end_column=product_col_idx)
            merged_cell = ws.cell(row=start_row, column=product_col_idx)
            merged_cell.alignment = Alignment(vertical='center', horizontal='center')

    # 保存 Excel 文件
    wb.save(output_file)
    print(f"已将 {target_date} 的数据导出到 {output_file}。")


if __name__ == "__main__":
    target_date = '20250530'
    # 输出文件路径改为 Excel 格式
    output_file = r'c:\Users\Top\Desktop\九章量化\exported_data.xlsx'
    export_data(target_date, output_file)